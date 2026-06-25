from __future__ import annotations

import sqlite3
from datetime import date as Date
from io import BytesIO


class InventoryService:
    def __init__(self, connection: sqlite3.Connection):
        self.connection = connection

    def add_supplier(self, name: str, contact_info: str = "") -> int:
        cursor = self.connection.execute(
            "INSERT INTO suppliers(name, contact_info) VALUES (?, ?)",
            (name.strip(), contact_info.strip()),
        )
        self.connection.commit()
        return int(cursor.lastrowid)

    def add_item(
        self,
        article: str,
        name: str,
        unit: str = "шт",
        min_stock: float = 0,
        lead_time_days: int = 14,
    ) -> int:
        cursor = self.connection.execute(
            """
            INSERT INTO items(article, name, unit, min_stock, lead_time_days)
            VALUES (?, ?, ?, ?, ?)
            """,
            (article.strip(), name.strip(), unit.strip(), min_stock, lead_time_days),
        )
        self.connection.commit()
        return int(cursor.lastrowid)

    def search_items(self, query: str, limit: int = 10) -> list[sqlite3.Row]:
        like_query = f"%{query.strip()}%"
        return self.connection.execute(
            """
            SELECT * FROM items
            WHERE is_active = 1 AND (article LIKE ? OR name LIKE ?)
            ORDER BY article
            LIMIT ?
            """,
            (like_query, like_query, limit),
        ).fetchall()

    def list_suppliers(self) -> list[sqlite3.Row]:
        return self.connection.execute(
            "SELECT * FROM suppliers WHERE is_active = 1 ORDER BY name"
        ).fetchall()

    def list_warehouses(self) -> list[sqlite3.Row]:
        return self.connection.execute("SELECT * FROM warehouses ORDER BY name").fetchall()

    def list_currencies(self) -> list[sqlite3.Row]:
        return self.connection.execute("SELECT * FROM currencies ORDER BY code").fetchall()

    def get_currency_id(self, code: str) -> int:
        row = self.connection.execute(
            "SELECT id FROM currencies WHERE code = ?",
            (code.upper(),),
        ).fetchone()
        if row is None:
            raise ValueError(f"Unknown currency: {code}")
        return int(row["id"])

    def record_receipt(
        self,
        item_id: int,
        qty: float,
        warehouse_id: int,
        supplier_id: int,
        price: float,
        currency_id: int,
        created_by: int,
        comment: str = "",
    ) -> None:
        self.connection.execute(
            """
            INSERT INTO operations(
                type, item_id, qty, warehouse_id, supplier_id, price,
                currency_id, comment, created_by
            ) VALUES ('receipt', ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (item_id, qty, warehouse_id, supplier_id, price, currency_id, comment, created_by),
        )
        self.connection.execute(
            """
            INSERT INTO supplier_prices(supplier_id, item_id, price, currency_id)
            VALUES (?, ?, ?, ?)
            """,
            (supplier_id, item_id, price, currency_id),
        )
        self.connection.commit()

    def record_consumption(
        self,
        item_id: int,
        qty: float,
        warehouse_id: int,
        created_by: int,
        comment: str = "",
    ) -> None:
        self.connection.execute(
            """
            INSERT INTO operations(type, item_id, qty, warehouse_id, comment, created_by)
            VALUES ('consumption', ?, ?, ?, ?, ?)
            """,
            (item_id, qty, warehouse_id, comment, created_by),
        )
        self.connection.commit()

    def stock_for_item_warehouse(self, item_id: int, warehouse_id: int) -> float:
        row = self.connection.execute(
            """
            SELECT COALESCE(SUM(CASE WHEN type = 'receipt' THEN qty ELSE -qty END), 0) AS qty
            FROM operations
            WHERE item_id = ? AND warehouse_id = ?
            """,
            (item_id, warehouse_id),
        ).fetchone()
        return float(row["qty"])

    def stock_rows(self, limit: int = 100) -> list[sqlite3.Row]:
        return self.connection.execute(
            """
            WITH movement AS (
                SELECT
                    item_id,
                    warehouse_id,
                    SUM(CASE WHEN type = 'receipt' THEN qty ELSE -qty END) AS qty
                FROM operations
                GROUP BY item_id, warehouse_id
            )
            SELECT
                i.id,
                i.article,
                i.name,
                i.unit,
                i.min_stock,
                COALESCE(SUM(CASE WHEN w.name = 'МСК' THEN m.qty ELSE 0 END), 0) AS msk_qty,
                COALESCE(SUM(CASE WHEN w.name = 'ТВЕРЬ' THEN m.qty ELSE 0 END), 0) AS tver_qty,
                COALESCE(SUM(m.qty), 0) AS total_qty,
                MAX(0, i.min_stock - COALESCE(SUM(m.qty), 0)) AS need_qty
            FROM items i
            LEFT JOIN movement m ON m.item_id = i.id
            LEFT JOIN warehouses w ON w.id = m.warehouse_id
            WHERE i.is_active = 1
            GROUP BY i.id
            ORDER BY need_qty DESC, i.article
            LIMIT ?
            """,
            (limit,),
        ).fetchall()

    def purchase_rows(self) -> list[sqlite3.Row]:
        return [row for row in self.stock_rows() if row["need_qty"] > 0]

    def get_state(self, key: str) -> str | None:
        row = self.connection.execute(
            "SELECT value FROM app_state WHERE key = ?", (key,)
        ).fetchone()
        return row["value"] if row else None

    def set_state(self, key: str, value: str) -> None:
        self.connection.execute(
            "INSERT OR REPLACE INTO app_state(key, value) VALUES (?, ?)",
            (key, value),
        )
        self.connection.commit()

    def history(
        self,
        item_id: int | None = None,
        days: int = 30,
        limit: int = 50,
    ) -> list[sqlite3.Row]:
        where = "WHERE o.created_at >= date('now', ?)"
        params: list = [f"-{days} days"]
        if item_id is not None:
            where += " AND o.item_id = ?"
            params.append(item_id)
        return self.connection.execute(
            f"""
            SELECT
                o.id, o.type, o.qty, o.price, o.comment, o.created_at,
                i.article, i.name, i.unit,
                w.name AS warehouse,
                s.name AS supplier,
                c.code AS currency
            FROM operations o
            JOIN items i ON i.id = o.item_id
            JOIN warehouses w ON w.id = o.warehouse_id
            LEFT JOIN suppliers s ON s.id = o.supplier_id
            LEFT JOIN currencies c ON c.id = o.currency_id
            {where}
            ORDER BY o.created_at DESC
            LIMIT ?
            """,
            (*params, limit),
        ).fetchall()

    def generate_excel(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()

        ws_stock = wb.active
        ws_stock.title = "Остатки"
        headers = ["Артикул", "Наименование", "Ед", "МСК", "ТВЕРЬ", "Итого", "Min", "К заказу"]
        ws_stock.append(headers)
        for row in self.stock_rows():
            ws_stock.append([
                row["article"], row["name"], row["unit"],
                row["msk_qty"], row["tver_qty"], row["total_qty"],
                row["min_stock"], row["need_qty"],
            ])

        ws_ops = wb.create_sheet("Движения")
        headers_ops = ["Дата", "Тип", "Артикул", "Наименование", "Кол-во", "Цена",
                        "Валюта", "Склад", "Поставщик", "Комментарий"]
        ws_ops.append(headers_ops)
        for row in self.history(days=365 * 3, limit=5000):
            ws_ops.append([
                row["created_at"],
                "Приход" if row["type"] == "receipt" else "Расход",
                row["article"], row["name"], row["qty"],
                row["price"], row["currency"],
                row["warehouse"], row["supplier"], row["comment"],
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
